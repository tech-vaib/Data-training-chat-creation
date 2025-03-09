import os
import pandas as pd
import numpy as np
import torch
import pickle
import logging
import re
from typing import List, Dict, Union, Optional, Tuple, Any
from pathlib import Path

# File processing libraries
import PyPDF2
from docx import Document
import openpyxl
import csv

# NLP and ML libraries
from transformers import AutoTokenizer, AutoModelForCausalLM, TrainingArguments, Trainer
from transformers import DataCollatorForLanguageModeling
from datasets import Dataset
from sklearn.model_selection import train_test_split
from transformers import GPT2LMHeadModel, GPT2Config
from torch.utils.data import DataLoader

# For UI
import gradio as gr

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Determine device for PyTorch
device = torch.device("cuda" if torch.cuda.is_available() else 
                     "mps" if torch.backends.mps.is_available() else 
                     "cpu")
logger.info(f"Using device: {device}")

class DocumentProcessor:
    """Process and extract text from various document formats"""
    
    def __init__(self, storage_dir: str = "data"):
        self.storage_dir = storage_dir
        os.makedirs(storage_dir, exist_ok=True)
        os.makedirs(os.path.join(storage_dir, "public"), exist_ok=True)
        os.makedirs(os.path.join(storage_dir, "private"), exist_ok=True)
        logger.info(f"Document processor initialized with storage in {storage_dir}")
    
    def save_file(self, file_path: str, file_content, is_public: bool = False) -> str:
        """Save uploaded file and return the path"""
        # Determine target directory based on privacy setting
        target_dir = "public" if is_public else "private"
        filename = os.path.basename(file_path)
        save_path = os.path.join(self.storage_dir, target_dir, filename)
        
        # Save the file
        with open(save_path, "wb") as f:
            f.write(file_content)
        
        logger.info(f"File saved to {save_path} with privacy setting: {'public' if is_public else 'private'}")
        return save_path
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from various file formats"""
        file_extension = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_extension == '.pdf':
                return self._extract_from_pdf(file_path)
            elif file_extension == '.txt':
                return self._extract_from_txt(file_path)
            elif file_extension == '.docx':
                return self._extract_from_docx(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return self._extract_from_excel(file_path)
            elif file_extension == '.csv':
                return self._extract_from_csv(file_path)
            else:
                error_msg = f"Unsupported file format: {file_extension}"
                logger.error(error_msg)
                raise ValueError(error_msg)
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            raise
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text = ""
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        logger.info(f"Extracted {len(text)} characters from PDF: {file_path}")
        return text
    
    def _extract_from_txt(self, file_path: str) -> str:
        """Extract text from TXT file"""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            text = file.read()
        logger.info(f"Extracted {len(text)} characters from TXT: {file_path}")
        return text
    
    def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        logger.info(f"Extracted {len(text)} characters from DOCX: {file_path}")
        return text
    
    def _extract_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"Sheet: {sheet}\n"
            
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                text += "\t".join(row_values) + "\n"
            
            text += "\n"
        
        logger.info(f"Extracted {len(text)} characters from Excel: {file_path}")
        return text
    
    def _extract_from_csv(self, file_path: str) -> str:
        """Extract text from CSV file"""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            reader = csv.reader(file)
            rows = list(reader)
            
        text = "\n".join(["\t".join(row) for row in rows])
        logger.info(f"Extracted {len(text)} characters from CSV: {file_path}")
        return text

class DataManager:
    """Manages data for model training"""
    
    def __init__(self, data_dir: str = "processed_data"):
        self.data_dir = data_dir
        os.makedirs(data_dir, exist_ok=True)
        self.documents = []
        self.document_metadata = []
        self.document_summaries = []  # Added to store document summaries
        logger.info(f"Data manager initialized with directory: {data_dir}")
    
    def add_document(self, text: str, metadata: Dict[str, Any]) -> None:
        """Add document and its metadata to the collection"""
        self.documents.append(text)
        self.document_metadata.append(metadata)
        
        # Generate and store a simple summary
        summary = self._generate_simple_summary(text)
        self.document_summaries.append(summary)
        
        logger.info(f"Added document with metadata: {metadata}")
    
    def _generate_simple_summary(self, text: str, max_length: int = 200) -> str:
        """Generate a simple summary of the text"""
        # Take first few sentences as a summary
        sentences = re.split(r'(?<=[.!?])\s+', text[:1000])
        summary = ' '.join(sentences[:3])
        
        if len(summary) > max_length:
            summary = summary[:max_length] + "..."
            
        return summary
    
    def prepare_for_training(self, chunk_size: int = 512, overlap: int = 50) -> List[str]:
        """Prepare documents for training by chunking them into appropriate sizes"""
        chunks = []
        
        for doc in self.documents:
            # Enhanced chunking strategy - tries to maintain sentence boundaries
            doc_chunks = self._smart_chunk_text(doc, chunk_size, overlap)
            chunks.extend(doc_chunks)
        
        logger.info(f"Prepared {len(chunks)} chunks for training")
        return chunks
    
    def _chunk_text(self, text: str, chunk_size: int, overlap: int) -> List[str]:
        """Split text into overlapping chunks of specified size"""
        words = text.split()
        chunks = []
        
        if len(words) <= chunk_size:
            return [text]
        
        i = 0
        while i < len(words):
            chunk = " ".join(words[i:i + chunk_size])
            chunks.append(chunk)
            i += chunk_size - overlap
        
        return chunks
    
    def _smart_chunk_text(self, text: str, chunk_size: int, overlap: int) -> List[str]:
        """Split text into overlapping chunks, trying to maintain sentence boundaries"""
        # Split text into sentences
        sentences = re.split(r'(?<=[.!?])\s+', text)
        chunks = []
        current_chunk = []
        current_length = 0
        
        for sentence in sentences:
            sentence_words = sentence.split()
            sentence_length = len(sentence_words)
            
            # If adding this sentence would exceed chunk size
            if current_length + sentence_length > chunk_size:
                # If we have something in the current chunk, add it
                if current_chunk:
                    chunks.append(" ".join(current_chunk))
                
                # Reset current chunk but keep some overlap
                if overlap > 0 and current_chunk:
                    # Take the last few words for overlap
                    overlap_words = min(overlap, len(current_chunk))
                    current_chunk = current_chunk[-overlap_words:]
                    current_length = len(current_chunk)
                else:
                    current_chunk = []
                    current_length = 0
            
            # Add the current sentence
            current_chunk.append(sentence)
            current_length += sentence_length
        
        # Add the last chunk if there's anything left
        if current_chunk:
            chunks.append(" ".join(current_chunk))
        
        return chunks
    
    def save_processed_data(self, filename: str = "processed_data.pkl") -> str:
        """Save processed data to file"""
        save_path = os.path.join(self.data_dir, filename)
        
        data = {
            "documents": self.documents,
            "metadata": self.document_metadata,
            "summaries": self.document_summaries
        }
        
        with open(save_path, 'wb') as f:
            pickle.dump(data, f)
        
        logger.info(f"Saved processed data to {save_path}")
        return save_path
    
    def load_processed_data(self, filename: str = "processed_data.pkl") -> None:
        """Load processed data from file"""
        load_path = os.path.join(self.data_dir, filename)
        
        with open(load_path, 'rb') as f:
            data = pickle.load(f)
        
        self.documents = data["documents"]
        self.document_metadata = data["metadata"]
        if "summaries" in data:
            self.document_summaries = data["summaries"]
        else:
            # Generate summaries if they don't exist in the loaded data
            self.document_summaries = [self._generate_simple_summary(doc) for doc in self.documents]
        
        logger.info(f"Loaded {len(self.documents)} documents from {load_path}")
    
    def get_document_contexts(self, query: str, top_n: int = 3) -> str:
        """Get relevant document contexts based on a query"""
        # Simple keyword-based relevance
        relevant_docs = []
        query_terms = set(query.lower().split())
        
        # Score each document based on query term matches
        doc_scores = []
        for i, doc in enumerate(self.documents):
            score = sum(1 for term in query_terms if term in doc.lower())
            doc_scores.append((i, score))
        
        # Sort by score and get top_n
        top_docs = sorted(doc_scores, key=lambda x: x[1], reverse=True)[:top_n]
        
        # Get the summaries of top docs
        context = "Related information from your documents:\n\n"
        for doc_idx, score in top_docs:
            if score > 0:  # Only include if there's some relevance
                context += f"- {self.document_summaries[doc_idx]}\n\n"
        
        return context if len(context) > 40 else ""  # Return empty if no relevant docs found

class ModelManager:
    """Manages model training, fine-tuning, and inference"""
    
 # Update the __init__ method in the ModelManager class

def __init__(self, model_name: str = "gpt2", model_dir: str = "model"):
    self.model_name = model_name
    self.model_dir = model_dir
    os.makedirs(model_dir, exist_ok=True)
    
    self.tokenizer = None
    self.model = None
    self.response_generation_config = {
        "temperature": 0.7,
        "top_p": 0.9,
        "top_k": 50,  # Added top_k parameter
        "repetition_penalty": 1.1,
        "no_repeat_ngram_size": 3,
        "max_length": 200
    }
    
    logger.info(f"Model manager initialized with model: {model_name}")
    
    def train_model_advanced(self, train_dataset: Dataset, val_dataset: Dataset, 
          output_dir: str = None, epochs: int = 3, 
          batch_size: int = 8, learning_rate: float = 5e-5,
          weight_decay: float = 0.01, warmup_ratio: float = 0.05,
          gradient_accumulation_steps: int = 4, save_steps: int = 400) -> None:
     """Train/fine-tune the model with advanced options"""
    if output_dir is None:
        output_dir = os.path.join(self.model_dir, "trained")
    
    # Data collator for language modeling
    data_collator = DataCollatorForLanguageModeling(
        tokenizer=self.tokenizer, mlm=False
    )
    
    # Calculate warmup steps based on ratio
    total_steps = len(train_dataset) // (batch_size * gradient_accumulation_steps) * epochs
    warmup_steps = int(total_steps * warmup_ratio)
    
    # Enhanced training arguments with advanced parameters
    training_args = TrainingArguments(
        output_dir=output_dir,
        overwrite_output_dir=True,
        num_train_epochs=epochs,
        per_device_train_batch_size=batch_size,
        per_device_eval_batch_size=batch_size,
        eval_steps=save_steps // 2,  # Evaluate more frequently than saving
        save_steps=save_steps,
        save_total_limit=2,  # Keep only the best 2 checkpoints
        warmup_steps=warmup_steps,
        learning_rate=learning_rate,
        weight_decay=weight_decay,  # Regularization parameter
        logging_dir=os.path.join(self.model_dir, "logs"),
        logging_steps=100,
        eval_strategy="steps",  # Use the corrected parameter name
        load_best_model_at_end=True,  # Load the best model at the end of training
        metric_for_best_model="loss",
        greater_is_better=False,
        gradient_accumulation_steps=gradient_accumulation_steps,  # For effective larger batch sizes
        fp16=torch.cuda.is_available(),  # Use mixed precision if GPU available
        # Additional optimization settings
        adam_beta1=0.9,
        adam_beta2=0.999,
        adam_epsilon=1e-8,
        max_grad_norm=1.0,  # Gradient clipping
        dataloader_num_workers=4,  # Parallel data loading
        dataloader_pin_memory=True,  # Speed up data transfer to GPU
    )
    
    # Initialize trainer
    trainer = Trainer(
        model=self.model,
        args=training_args,
        data_collator=data_collator,
        train_dataset=train_dataset,
        eval_dataset=val_dataset,
    )
    
    # Log training configuration
    logger.info(f"Starting model training with the following configuration:")
    logger.info(f"- Epochs: {epochs}")
    logger.info(f"- Batch size: {batch_size}")
    logger.info(f"- Learning rate: {learning_rate}")
    logger.info(f"- Weight decay: {weight_decay}")
    logger.info(f"- Warmup steps: {warmup_steps} ({warmup_ratio} of total steps)")
    logger.info(f"- Gradient accumulation steps: {gradient_accumulation_steps}")
    logger.info(f"- Save steps: {save_steps}")
    
    # Train the model
    logger.info(f"Starting model training for {epochs} epochs...")
    trainer.train()
    
    # Save the model and tokenizer
    trainer.save_model(output_dir)
    self.tokenizer.save_pretrained(output_dir)
    
    logger.info(f"Model training completed and saved to {output_dir}")

    def update_generation_config(self, config: Dict[str, Any]) -> None:
        """Update response generation configuration"""
        self.response_generation_config.update(config)
        logger.info(f"Updated generation config: {self.response_generation_config}")
    
    def initialize_model(self) -> None:
        """Initialize the model and tokenizer"""
        try:
            self.tokenizer = AutoTokenizer.from_pretrained(self.model_name)
            self.model = AutoModelForCausalLM.from_pretrained(self.model_name).to(device)
            
            # Ensure the tokenizer has a padding token
            if self.tokenizer.pad_token is None:
                self.tokenizer.pad_token = self.tokenizer.eos_token
                
            # Make sure pad_token_id is set in the model config
            if self.model.config.pad_token_id is None:
                self.model.config.pad_token_id = self.tokenizer.pad_token_id
            
            logger.info(f"Initialized model and tokenizer: {self.model_name} on {device}")
        except Exception as e:
            logger.error(f"Error initializing model: {str(e)}")
            raise
    
    def prepare_dataset(self, texts: List[str]) -> Tuple[Dataset, Dataset]:
        """Prepare dataset for training"""
        if self.tokenizer is None:
            self.initialize_model()
        
        # Tokenize texts
        encodings = self.tokenizer(texts, truncation=True, padding="max_length", 
                                   max_length=512, return_special_tokens_mask=True)
        
        # Create dataset
        dataset = Dataset.from_dict(encodings)
        
        # Split into train and validation
        train_val = dataset.train_test_split(test_size=0.1)
        train_dataset = train_val['train']
        val_dataset = train_val['test']
        
        logger.info(f"Prepared dataset with {len(train_dataset)} training samples "
                   f"and {len(val_dataset)} validation samples")
        
        return train_dataset, val_dataset
    
    def train(self, train_dataset: Dataset, val_dataset: Dataset, 
              output_dir: str = None, epochs: int = 3, 
              batch_size: int = 8, learning_rate: float = 5e-5) -> None:
        """Train/fine-tune the model"""
        if output_dir is None:
            output_dir = os.path.join(self.model_dir, "trained")
        
        # Data collator for language modeling
        data_collator = DataCollatorForLanguageModeling(
            tokenizer=self.tokenizer, mlm=False
        )
        
        # Enhanced training arguments
        training_args = TrainingArguments(
            output_dir=output_dir,
            overwrite_output_dir=True,
            num_train_epochs=epochs,
            per_device_train_batch_size=batch_size,
            per_device_eval_batch_size=batch_size,
            eval_steps=200,  # Evaluate more frequently
            save_steps=400,
            save_total_limit=2,  # Keep only the best 2 checkpoints
            warmup_steps=100,
            learning_rate=learning_rate,
            weight_decay=0.01,  # Add weight decay for regularization
            logging_dir=os.path.join(self.model_dir, "logs"),
            logging_steps=100,
            evaluation_strategy="steps",
            load_best_model_at_end=True,  # Load the best model at the end of training
            metric_for_best_model="loss",
            greater_is_better=False,
            gradient_accumulation_steps=4,  # For effective larger batch sizes
            fp16=torch.cuda.is_available(),  # Use mixed precision if GPU available
        )
        
        # Initialize trainer
        trainer = Trainer(
            model=self.model,
            args=training_args,
            data_collator=data_collator,
            train_dataset=train_dataset,
            eval_dataset=val_dataset,
        )
        
        # Train the model
        logger.info(f"Starting model training for {epochs} epochs...")
        trainer.train()
        
        # Save the model and tokenizer
        trainer.save_model(output_dir)
        self.tokenizer.save_pretrained(output_dir)
        
        logger.info(f"Model training completed and saved to {output_dir}")
    
    def load_trained_model(self, model_path: str) -> None:
        """Load a trained model"""
        try:
            self.tokenizer = AutoTokenizer.from_pretrained(model_path)
            self.model = AutoModelForCausalLM.from_pretrained(model_path).to(device)
            
            # Ensure the tokenizer has a padding token
            if self.tokenizer.pad_token is None:
                self.tokenizer.pad_token = self.tokenizer.eos_token
                
            # Make sure pad_token_id is set in the model config
            if self.model.config.pad_token_id is None:
                self.model.config.pad_token_id = self.tokenizer.pad_token_id
            
            logger.info(f"Loaded trained model from {model_path} to {device}")
        except Exception as e:
            logger.error(f"Error loading trained model: {str(e)}")
            raise
    
    def generate_response(self, prompt: str, max_length: int = None) -> str:
       """Generate a response using the trained model with enhanced parameters"""
    if self.model is None or self.tokenizer is None:
        raise ValueError("Model and tokenizer must be initialized before generating responses")
    
    # Use provided max_length or default from config
    if max_length is None:
        max_length = self.response_generation_config["max_length"]
    
    try:
        # Encode the prompt and send to the correct device
        inputs = self.tokenizer(prompt, return_tensors="pt", padding=True)
        inputs = {k: v.to(device) for k, v in inputs.items()}
        
        # Create an attention mask if it's not present
        if 'attention_mask' not in inputs:
            inputs['attention_mask'] = torch.ones_like(inputs['input_ids'])
        
        # Get top_k if it exists in the config, otherwise use default
        top_k = self.response_generation_config.get("top_k", 50)
        
        # Generate response with enhanced parameters
        outputs = self.model.generate(
            input_ids=inputs['input_ids'],
            attention_mask=inputs['attention_mask'],
            max_length=max_length + inputs['input_ids'].shape[1],
            num_return_sequences=1,
            pad_token_id=self.tokenizer.pad_token_id,
            do_sample=True,
            temperature=self.response_generation_config["temperature"],
            top_p=self.response_generation_config["top_p"],
            top_k=top_k,  # Add top_k sampling
            repetition_penalty=self.response_generation_config["repetition_penalty"],
            no_repeat_ngram_size=self.response_generation_config["no_repeat_ngram_size"],
            early_stopping=False,  # Disable early stopping when not using beam search
            # Additional parameters to improve quality
            min_length=inputs['input_ids'].shape[1] + 10,  # Ensure responses have some length
            length_penalty=1.0,  # Neutral length penalty
            num_beam_groups=1,
            diversity_penalty=0.0,
        )
        
        # Move output back to CPU before decoding
        outputs = outputs.cpu()
        
        # Decode the response
        generated_text = self.tokenizer.decode(outputs[0], skip_special_tokens=True)
        
        # Extract only the newly generated text (remove the prompt)
        response = generated_text[len(prompt):].strip()
        
        # Post-process to clean up the response
        response = self._post_process_response(response)
        
        logger.info(f"Generated response of length {len(response)}")
        return response
    except Exception as e:
        logger.error(f"Error generating response: {str(e)}")
        raise

    def _post_process_response(self, response: str) -> str:
        """Clean up the generated response"""
        # Remove any "User:" or similar markers that might have been generated
        response = re.sub(r'User:|Assistant:', '', response)
        
        # Truncate at any newline followed by "User:" or similar (stopping at turns)
        match = re.search(r'\n(User:|Human:)', response)
        if match:
            response = response[:match.start()]
        
        # Remove excess whitespace
        response = re.sub(r'\s+', ' ', response).strip()
        
        return response

class ChatInterface:
    """Interface for interacting with the trained model"""
    
    def __init__(self, document_processor: DocumentProcessor, 
                 data_manager: DataManager, model_manager: ModelManager):
        self.document_processor = document_processor
        self.data_manager = data_manager
        self.model_manager = model_manager
        self.conversation_history = []
        self.max_history_turns = 5  # Store last 5 conversation turns
        self.system_prompt = """You are a helpful AI assistant trained on custom data. 
        Answer questions based on the information provided in the documents.
        If you don't know the answer, say so rather than making something up."""
        logger.info("Chat interface initialized")
    
def train_model_advanced(self, epochs: int = 8, batch_size: int = 8, 
                        learning_rate: float = 5e-5, weight_decay: float = 0.01,
                        warmup_ratio: float = 0.05, gradient_accumulation_steps: int = 4,
                        save_steps: int = 400) -> str:
    """Train the model using the uploaded files with advanced options"""
    try:
        # Prepare data
        chunks = self.data_manager.prepare_for_training()
        if not chunks:
            return "No data available for training. Please upload some files first."
        
        # Save processed data
        self.data_manager.save_processed_data()
        
        # Initialize model
        self.model_manager.initialize_model()
        
        # Prepare dataset
        train_dataset, val_dataset = self.model_manager.prepare_dataset(chunks)
        
        # Train model with advanced parameters
        self.model_manager.train_model_advanced(
            train_dataset=train_dataset,
            val_dataset=val_dataset,
            epochs=epochs,
            batch_size=batch_size,
            learning_rate=learning_rate,
            weight_decay=weight_decay,
            warmup_ratio=warmup_ratio,
            gradient_accumulation_steps=gradient_accumulation_steps,
            save_steps=save_steps
        )
        
        # Also update the model's generation with the latest parameters
        self.model_manager.generate_response("Test prompt", max_length=50)  # Force loading
        
        return "Model training completed successfully with advanced parameters!"
    except Exception as e:
        error_msg = f"Error training model: {str(e)}"
        logger.error(error_msg)
        return error_msg    
    
    def upload_file(self, file_path: str, file_content, is_public: bool = False) -> str:
        """Upload and process a file"""
        try:
            # Save the file
            saved_path = self.document_processor.save_file(file_path, file_content, is_public)
            
            # Extract text
            text = self.document_processor.extract_text(saved_path)
            
            # Add to data manager
            metadata = {
                "filename": os.path.basename(file_path),
                "is_public": is_public,
                "path": saved_path,
                "length": len(text),
                "timestamp": pd.Timestamp.now().isoformat()
            }
            self.data_manager.add_document(text, metadata)
            
            return f"Successfully uploaded and processed {os.path.basename(file_path)}"
        except Exception as e:
            error_msg = f"Error processing file {file_path}: {str(e)}"
            logger.error(error_msg)
            return error_msg
    
    def train_model(self, epochs: int = 8, batch_size: int = 8) -> str:
        """Train the model using the uploaded files"""
        try:
            # Prepare data
            chunks = self.data_manager.prepare_for_training()
            if not chunks:
                return "No data available for training. Please upload some files first."
            
            # Save processed data
            self.data_manager.save_processed_data()
            
            # Initialize model
            self.model_manager.initialize_model()
            
            # Prepare dataset
            train_dataset, val_dataset = self.model_manager.prepare_dataset(chunks)
            
            # Train model
            self.model_manager.train(
                train_dataset=train_dataset,
                val_dataset=val_dataset,
                epochs=epochs,
                batch_size=batch_size
            )
            
            return "Model training completed successfully!"
        except Exception as e:
            error_msg = f"Error training model: {str(e)}"
            logger.error(error_msg)
            return error_msg
    
    def get_response(self, message: str) -> str:
        """Get response from the model based on user message"""
        try:
            if self.model_manager.model is None:
                return "Model is not trained yet. Please train the model first."
            
            # Get relevant document context
            doc_context = self.data_manager.get_document_contexts(message)
            
            # Build conversation history
            history_text = ""
            for turn in self.conversation_history[-self.max_history_turns:]:
                history_text += f"User: {turn[0]}\nAssistant: {turn[1]}\n\n"
            
            # Create prompt with system prompt, context, history, and message
            prompt = f"{self.system_prompt}\n\n"
            if doc_context:
                prompt += f"{doc_context}\n\n"
            if history_text:
                prompt += f"{history_text}"
            prompt += f"User: {message}\nAssistant:"
            
            # Generate response
            response = self.model_manager.generate_response(prompt)
            
            # Add to conversation history
            self.conversation_history.append((message, response))
            
            return response
        except Exception as e:
            error_msg = f"Error generating response: {str(e)}"
            logger.error(error_msg)
            return error_msg
    
    def set_system_prompt(self, prompt: str) -> None:
        """Set a custom system prompt"""
        self.system_prompt = prompt
        logger.info(f"Updated system prompt: {prompt[:50]}...")
    
    def clear_conversation(self) -> None:
        """Clear the conversation history"""
        self.conversation_history = []
        logger.info("Conversation history cleared")

# Create the Gradio UI
def create_ui():
    """Create Gradio UI for the application"""
    # Initialize components
    document_processor = DocumentProcessor()
    data_manager = DataManager()
    model_manager = ModelManager()
    chat_interface = ChatInterface(document_processor, data_manager, model_manager)
    
    # Define file upload function
    def upload_func(files, is_public):
        results = []
        for file in files:
            with open(file.name, "rb") as f:
                content = f.read()
            result = chat_interface.upload_file(file.name, content, is_public)
            results.append(result)
        return "\n".join(results)
    
    # Enhanced training function with more parameters
    def train_func(epochs, batch_size, learning_rate, weight_decay, warmup_ratio, 
                  temperature, top_p, top_k, repetition_penalty, max_length, 
                  gradient_accumulation_steps, save_steps):
        # Update generation config
        model_manager.update_generation_config({
            "temperature": temperature,
            "top_p": top_p,
            "top_k": top_k,
            "repetition_penalty": repetition_penalty,
            "max_length": max_length
        })
        
        # Return training result with all parameters to track settings
        result = chat_interface.train_model_advanced(
            epochs=int(epochs), 
            batch_size=int(batch_size),
            learning_rate=learning_rate,
            weight_decay=weight_decay,
            warmup_ratio=warmup_ratio,
            gradient_accumulation_steps=int(gradient_accumulation_steps),
            save_steps=int(save_steps)
        )
        
        # Return the result along with a summary of parameters used
        param_summary = f"Training completed with parameters:\n" + \
                        f"• Epochs: {epochs}\n" + \
                        f"• Batch Size: {batch_size}\n" + \
                        f"• Learning Rate: {learning_rate}\n" + \
                        f"• Weight Decay: {weight_decay}\n" + \
                        f"• Generation Temp: {temperature}\n" + \
                        f"• Top-p: {top_p}\n" + \
                        f"• Top-k: {top_k}\n" + \
                        f"• Repetition Penalty: {repetition_penalty}\n"
        
        return f"{result}\n\n{param_summary}"
    
    # Define chat function
    def chat_func(message, history):
        bot_message = chat_interface.get_response(message)
        history.append((message, bot_message))
        return "", history
    
    # Define system prompt update function
    def update_system_prompt(prompt):
        chat_interface.set_system_prompt(prompt)
        return f"System prompt updated successfully! ({len(prompt)} characters)"
    
    # Define function to clear chat
    def clear_chat():
        chat_interface.clear_conversation()
        return []
    
    # Build the UI
    with gr.Blocks(title="Custom NLP Model Training") as demo:
        gr.Markdown("# Custom NLP Model Training and Chat Interface")
        
        with gr.Tab("Upload and Process Files"):
            with gr.Row():
                files_input = gr.File(file_count="multiple", label="Upload Files (PDF, TXT, DOCX, Excel, CSV)")
                is_public = gr.Checkbox(label="Make files public", value=True)
            
            upload_button = gr.Button("Upload and Process Files")
            upload_output = gr.Textbox(label="Upload Results")
            
            upload_button.click(
                fn=upload_func,
                inputs=[files_input, is_public],
                outputs=upload_output
            )
        
        with gr.Tab("Train Model"):
            gr.Markdown("### Training Parameters")
            
            with gr.Accordion("Basic Parameters", open=True):
                with gr.Row():
                    epochs_input = gr.Slider(minimum=1, maximum=50, value=8, step=1, 
                                            label="Training Epochs")
                    batch_size_input = gr.Slider(minimum=1, maximum=32, value=8, step=1, 
                                                label="Batch Size")
                
                with gr.Row():
                    learning_rate_input = gr.Slider(minimum=1e-6, maximum=1e-3, value=5e-5, 
                                                  label="Learning Rate")
                    weight_decay_input = gr.Slider(minimum=0, maximum=0.1, value=0.01, step=0.005, 
                                                 label="Weight Decay")
            
            with gr.Accordion("Advanced Training Parameters", open=False):
                with gr.Row():
                    warmup_ratio_input = gr.Slider(minimum=0, maximum=0.2, value=0.05, step=0.01, 
                                                 label="Warmup Ratio")
                    gradient_accum_input = gr.Slider(minimum=1, maximum=16, value=4, step=1, 
                                                   label="Gradient Accumulation Steps")
                
                with gr.Row():
                    save_steps_input = gr.Slider(minimum=100, maximum=1000, value=400, step=100, 
                                               label="Save Steps")
            
            gr.Markdown("### Response Generation Parameters")
            with gr.Accordion("Text Generation Settings", open=True):
                with gr.Row():
                    temperature_input = gr.Slider(minimum=0.1, maximum=2.0, value=0.7, step=0.1, 
                                                label="Temperature")
                    top_p_input = gr.Slider(minimum=0.1, maximum=1.0, value=0.9, step=0.05, 
                                          label="Top-p (nucleus sampling)")
                
                with gr.Row():
                    top_k_input = gr.Slider(minimum=1, maximum=100, value=50, step=1, 
                                          label="Top-k")
                    repetition_penalty_input = gr.Slider(minimum=1.0, maximum=2.0, value=1.1, step=0.1, 
                                                       label="Repetition Penalty")
                
                with gr.Row():
                    max_length_input = gr.Slider(minimum=50, maximum=500, value=200, step=10, 
                                               label="Max Response Length")
            
            train_button = gr.Button("Train Model", variant="primary")
            train_output = gr.Textbox(label="Training Results", lines=10)
            
            train_button.click(
                fn=train_func,
                inputs=[
                    epochs_input, batch_size_input, learning_rate_input, weight_decay_input, 
                    warmup_ratio_input, temperature_input, top_p_input, top_k_input, 
                    repetition_penalty_input, max_length_input, gradient_accum_input, 
                    save_steps_input
                ],
                outputs=train_output
            )
        
        with gr.Tab("Chat Interface"):
            with gr.Row():
                with gr.Column(scale=3):
                    system_prompt_input = gr.Textbox(
                        value=chat_interface.system_prompt,
                        label="System Prompt",
                        lines=3
                    )
                with gr.Column(scale=1):
                    update_prompt_button = gr.Button("Update System Prompt")
                    prompt_update_output = gr.Textbox(label="Update Status")
            
            chatbot = gr.Chatbot(label="Conversation")
            
            with gr.Row():
                msg = gr.Textbox(label="Message", placeholder="Type your message here...")
                clear_button = gr.Button("Clear Chat")
            
            msg.submit(
                fn=chat_func,
                inputs=[msg, chatbot],
                outputs=[msg, chatbot]
            )
            
            update_prompt_button.click(
                fn=update_system_prompt,
                inputs=system_prompt_input,
                outputs=prompt_update_output
            )
            
            clear_button.click(
                fn=clear_chat,
                inputs=[],
                outputs=chatbot
            )
    
    return demo

# Main entry point
if __name__ == "__main__":
    demo = create_ui()
    demo.launch(share=True)  # Added share=True for easier sharing